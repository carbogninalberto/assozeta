"""
@ copyright: Bakney SRL
"""
import hashlib
import hmac
import json
import logging
import re
from datetime import datetime

from bs4 import BeautifulSoup
from django.conf import settings
from django.http import HttpResponse
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.exceptions import ValidationError
from rest_framework.response import Response
from openpyxl import Workbook

from application.models import BillingSubscription, MedicalCertificate, Payment
from application.models.balance_sheet_models import CustomAccounts
from application.models.courses_models import Course, CourseSubscription, CourseSubscriptionInstallment
from application.models.invoices_models import Invoice
from application.models.payment_models import PaymentCategory, SupplierAndCustomers
from application.models.subscriptions_models import Subscription
from application.models.user_models import SportAssociation, SportAssociationMembershipCardConfiguration, User, Associate, SportAssociationInvoices, \
    Testimonial, Instructor
from application.serializers.auth_serializers import UserAuthSerializer, SportAssociationSerializer, \
    UserAuthUpdateSerializer, UserSettingsSerializer, UserTablesSettingsSerializer, SportAssociationToolSerializer, \
    SportAssociationInvoiceSerializer
from application.utils.api_utils import ApiMessages, is_valid_uuid, BalanceSheetData, compress_base64
from core.middleware import IsAuthenticated

logger = logging.getLogger(__name__)


@api_view(['PATCH'])
@permission_classes([IsAuthenticated])
def profile_update(request):
    logger.info("Updating profile", extra={'user_id': str(request.user.user_id)})

    user_data = UserAuthUpdateSerializer(data=request.data['user_data'])
    user = request.user if request.collaborator is None or request.collaborator is False else request.original_user

    if user_data.initial_data['first_name'] is not None:
        user.first_name = user_data.initial_data['first_name']

    if user_data.initial_data['last_name'] is not None:
        user.last_name = user_data.initial_data['last_name']

    if user_data.initial_data['username'] is not None:
        user.username = user_data.initial_data['username'].upper()

    if user_data.initial_data['avatar_image'] is not None:
        user.avatar_image = compress_base64(user_data.initial_data['avatar_image'], max_size=(512, 512))
    user.save()

    if user_data.initial_data['sport_association'] is not None and user.role == User.ASSOCIATION:
        sport_association = SportAssociation.objects.get(user=user)
        if user_data.initial_data['sport_association']['denomination'] is not None:
            sport_association.denomination = user_data.initial_data['sport_association']['denomination']

        if user_data.initial_data['sport_association']['tax_code'] is not None:
            sport_association.tax_code = user_data.initial_data['sport_association']['tax_code']

        if user_data.initial_data['sport_association']['address'] is not None:
            sport_association.address = user_data.initial_data['sport_association']['address']

        if user_data.initial_data['sport_association']['address_cap'] is not None:
            sport_association.address_cap = user_data.initial_data['sport_association']['address_cap']

        if user_data.initial_data['sport_association']['address_city'] is not None:
            sport_association.address_city = user_data.initial_data['sport_association']['address_city']

        if user_data.initial_data['sport_association']['document_header'] is not None:
            sport_association.document_header = user_data.initial_data['sport_association']['document_header']

        if user_data.initial_data['sport_association']['invoice_footer'] is not None:
            sport_association.invoice_footer = user_data.initial_data['sport_association']['invoice_footer']

        if user_data.initial_data['sport_association']['enable_quotes_management'] is not None:
            sport_association.enable_quotes_management = user_data.initial_data['sport_association']['enable_quotes_management']

        if user_data.initial_data['sport_association']['configuration'] is not None:
            sport_association.configuration = user_data.initial_data['sport_association']['configuration']

        if user_data.initial_data['sport_association']['federation'] is not None:
            sport_association.federation = user_data.initial_data['sport_association']['federation']

        if user_data.initial_data['sport_association']['enroll_number'] is not None:
            sport_association.enroll_number = user_data.initial_data['sport_association']['enroll_number']

        if user_data.initial_data['sport_association']['sport'] is not None:
            sport_association.sport = user_data.initial_data['sport_association']['sport']

        if user_data.initial_data['sport_association']['president_signature'] is not None:
            sport_association.president_signature = user_data.initial_data['sport_association']['president_signature']

        if user_data.initial_data['sport_association']['stamp'] is not None:
            sport_association.stamp = user_data.initial_data['sport_association']['stamp']

        if user_data.initial_data['sport_association']['president_first_name'] is not None:
            sport_association.president_first_name = user_data.initial_data['sport_association']['president_first_name']

        if user_data.initial_data['sport_association']['president_last_name'] is not None:
            sport_association.president_last_name = user_data.initial_data['sport_association']['president_last_name']

        if user_data.initial_data['sport_association']['stripe_available_methods'] is not None:
            sport_association.stripe_available_methods = user_data.initial_data['sport_association'][
                'stripe_available_methods']

        if user_data.initial_data['sport_association']['invoice_template'] is not None:
            sport_association.invoice_template = user_data.initial_data['sport_association'][
                'invoice_template']

        if user_data.initial_data['sport_association']['subscription_template'] is not None:
            sport_association.subscription_template = user_data.initial_data['sport_association'][
                'subscription_template']

        if user_data.initial_data['sport_association']['extra_text_invoices'] is not None:
            sport_association.extra_text_invoices = user_data.initial_data['sport_association'][
                'extra_text_invoices']

        if 'iban' in user_data.initial_data['sport_association'] and \
                user_data.initial_data['sport_association']['iban'] is not None:
            sport_association.iban = user_data.initial_data['sport_association']['iban']

        if 'abbreviated' in user_data.initial_data['sport_association'] and \
                user_data.initial_data['sport_association']['abbreviated'] is not None:
            sport_association.abbreviated = user_data.initial_data['sport_association']['abbreviated']

        if 'vat_number' in user_data.initial_data['sport_association'] and \
                user_data.initial_data['sport_association']['vat_number'] is not None:
            sport_association.vat_number = user_data.initial_data['sport_association']['vat_number']

        if 'website' in user_data.initial_data['sport_association'] and \
                user_data.initial_data['sport_association']['website'] is not None:
            sport_association.website = user_data.initial_data['sport_association']['website']

        if 'whatsapp' in user_data.initial_data['sport_association'] and \
                user_data.initial_data['sport_association']['whatsapp'] is not None:
            sport_association.whatsapp = user_data.initial_data['sport_association']['whatsapp']

        sport_association.checkout_info = user_data.initial_data['sport_association']['checkout_info']

        logger.info("Saving sport association profile updates", extra={'user_id': str(user.user_id), 'sport_association_id': str(sport_association.sport_association_id)})
        sport_association.save()

    logger.info("Profile updated successfully", extra={'user_id': str(user.user_id)})
    content = {}
    # logger.info(content)
    content['role'] = User.ROLE_CHOICES[user.role - 1][1]
    content['user_data'] = UserAuthSerializer(user).data
    # compress avatar image
    if content['user_data']['avatar_image'] is not None:
        content['user_data']['avatar_image'] = compress_base64(content['user_data']['avatar_image'])

    if user.role == User.ASSOCIATION:
        sport_association = SportAssociation.objects.get(user=user)
        content['user_data']['sport_association'] = SportAssociationSerializer(sport_association).data
        # compress logo image
        if content['user_data']['sport_association']['logo'] is not None:
            content['user_data']['sport_association']['logo'] = compress_base64(
                content['user_data']['sport_association']['logo'])

        content['user_data']['preview_and_custom_features'] = user.get_preveiw_and_custom_features()

    return Response(content, status=status.HTTP_200_OK)


@api_view(['PATCH'])
@permission_classes([IsAuthenticated])
def profile_update_subscription_template(request):

    sport_association_data = UserAuthUpdateSerializer(data=request.data['sport_association'])
    user = request.user

    if sport_association_data.initial_data is not None:
        sport_association = SportAssociation.objects.get(user=user)
        # update additional_sections
        if sport_association_data.initial_data['additional_sections'] is not None:
            # loop through additional_sections and remove if json field name or text is empty or if there are different
            # json field names with the same name
            additional_sections = sport_association_data.initial_data['additional_sections']

            if len(additional_sections) > 0:
                sport_association.additional_sections = []
                for section in additional_sections:
                    if 'name' in section.keys() and 'text' in section.keys():
                        if section['name'] == '' and len(section['name']) <= 3:
                            continue
                        if section['text'] == '' and len(section['text']) <= 3:
                            continue
                        section['text'] = BeautifulSoup(section['text'], features="html.parser").prettify()
                        section['name'] = section['name'].upper()
                        sport_association.additional_sections.append(section)
            else:
                sport_association.additional_sections = []

        if sport_association_data.initial_data['regulation'] is not None:
            sport_association.regulation = BeautifulSoup(sport_association_data.initial_data['regulation'],
            features="html.parser").\
                prettify()

        if sport_association_data.initial_data['demand'] is not None:
            sport_association.demand = BeautifulSoup(sport_association_data.initial_data['demand'],
            features="html.parser").prettify()

        # add the visibility settings
        if sport_association_data.initial_data['show_demand_to_athletes'] is not None:
            sport_association.show_demand_to_athletes = sport_association_data.initial_data['show_demand_to_athletes']
        if sport_association_data.initial_data['show_demand_to_members'] is not None:
            sport_association.show_demand_to_members = sport_association_data.initial_data['show_demand_to_members']
        if sport_association_data.initial_data['show_demand_to_both'] is not None:
            sport_association.show_demand_to_both = sport_association_data.initial_data['show_demand_to_both']
        if sport_association_data.initial_data['show_regulation_to_athletes'] is not None:
            sport_association.show_regulation_to_athletes = sport_association_data.initial_data['show_regulation_to_athletes']
        if sport_association_data.initial_data['show_regulation_to_members'] is not None:
            sport_association.show_regulation_to_members = sport_association_data.initial_data['show_regulation_to_members']
        if sport_association_data.initial_data['show_regulation_to_both'] is not None:
            sport_association.show_regulation_to_both = sport_association_data.initial_data['show_regulation_to_both']

        if sport_association_data.initial_data['logo'] is not None:
            sport_association.logo = compress_base64(
                sport_association_data.initial_data['logo'],
                max_size=(590, 130)
            )
        elif sport_association_data.initial_data['logo'] == '' or sport_association_data.initial_data['logo'] is None:
            sport_association.logo = None

        if sport_association_data.initial_data['subscription_fee'] is not None:
            fee = float(sport_association_data.initial_data['subscription_fee'])
            if fee >= 0:
                sport_association.subscription_fee = sport_association_data.initial_data['subscription_fee']
        if sport_association_data.initial_data['multiple_subscription_fee'] is not None:
            sport_association.multiple_subscription_fee = sport_association_data.initial_data['multiple_subscription_fee']
        if sport_association_data.initial_data['subscription_fee_plans'] is not None:
            plans = sport_association_data.initial_data['subscription_fee_plans']
            if len(plans) > 0:
                sport_association.subscription_fee_plans = []
                for plan in plans:
                    if plan['subscription_fee'] == '':
                        plan['subscription_fee'] = 0
                    if plan['name'] is not None and plan['subscription_fee'] is not None and float(plan['subscription_fee']) >= 0:
                        plan['subscription_fee'] = float(plan['subscription_fee'])
                        sport_association.subscription_fee_plans.append(plan)
            else:
                # should not allow empty plans and multiple subscription fee active
                sport_association.multiple_subscription_fee = False

        if sport_association_data.initial_data['membership_fee'] is not None:
            if sport_association_data.initial_data['membership_fee'] == '':
                sport_association.membership_fee = None
            else:
                fee = float(sport_association_data.initial_data['membership_fee'])
                if fee >= 0:
                    sport_association.membership_fee = sport_association_data.initial_data['membership_fee']
        if sport_association_data.initial_data['multiple_membership_fee'] is not None:
            sport_association.multiple_membership_fee = sport_association_data.initial_data[
                'multiple_membership_fee']
        if sport_association_data.initial_data['membership_fee_plans'] is not None:
            plans = sport_association_data.initial_data['membership_fee_plans']
            if len(plans) > 0:
                sport_association.membership_fee_plans = []
                for plan in plans:
                    if plan['membership_fee'] == '':
                        plan['membership_fee'] = 0
                    if plan['name'] is not None and plan['membership_fee'] is not None and float(
                            plan['membership_fee']) >= 0:
                        plan['membership_fee'] = float(plan['membership_fee'])
                        sport_association.membership_fee_plans.append(plan)
            else:
                # should not allow empty plans and multiple membership fee active
                sport_association.multiple_membership_fee = False
                sport_association.membership_fee_plans = []

        if sport_association_data.initial_data['enable_quotes_management'] is not None:
            sport_association.enable_quotes_management = sport_association_data.initial_data[
                'enable_quotes_management']

        sport_association.enabled_for = sport_association_data.initial_data['enabled_for']
        sport_association.additional_fields = sport_association_data.initial_data['additional_fields']
        sport_association.stripe_available_methods = sport_association_data.initial_data['stripe_available_methods']
        sport_association.invoice_template = sport_association_data.initial_data['invoice_template']
        sport_association.subscription_template = sport_association_data.initial_data['subscription_template']
        sport_association.extra_text_invoices = sport_association_data.initial_data['extra_text_invoices']

        sport_association.save()

    content = {}
    # logger.info(content)
    content['role'] = User.ROLE_CHOICES[user.role - 1][1]
    content['user_data'] = UserAuthSerializer(user).data
    content['user_data']['preview_and_custom_features'] = user.get_preveiw_and_custom_features()
    # compress avatar image
    if content['user_data']['avatar_image'] is not None:
        content['user_data']['avatar_image'] = compress_base64(content['user_data']['avatar_image'])

    if user.role == User.ASSOCIATION:
        sport_association = SportAssociation.objects.get(user=user)
        content['user_data']['sport_association'] = SportAssociationSerializer(sport_association).data
        # compress logo image
        if content['user_data']['sport_association']['logo'] is not None:
            content['user_data']['sport_association']['logo'] = compress_base64(
                content['user_data']['sport_association']['logo'],
                max_size=(590, 130)
            )

    return Response(content, status=status.HTTP_200_OK)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def profile_associates_course(request, uid):
    """
    Select all the associates that are subscribed on the course and are connected to the user,
    plus return a list of possible subscriptions
    """
    is_valid_uuid(uid)

    user = request.user
    sport_association = Course.objects.filter(course_id=uid).prefetch_related('sport_association').first().sport_association
    associates = Associate.objects.filter(user__in=[user, sport_association.user])
    subscribed_associates = CourseSubscription.objects.filter(course_id=uid)\
        .prefetch_related('subscription', 'subscription__associate').filter(subscription__associate__in=associates)\
        .values_list('subscription__associate__associate_id', flat=True)

    associates_subscribed = []
    associates_unsubscribed = []

    subscriptions = Subscription.objects.filter(
        user=user,
        associate__in=associates,
        sport_association=sport_association
    ).prefetch_related('associate')
    # make a flat list of subscriptions with the subscription_id and associate_id
    subscriptions = subscriptions.values_list('subscription_id', 'associate_id')

    for associate in associates.iterator():
        if associate.associate_id in subscribed_associates:
            associates_subscribed.append({
                'id': associate.associate_id,
                'full_name': associate.get_full_name(),
                'is_minor': associate.is_minor
            })
        else:
            # check if associate_id is in subscriptions
            # next() returns the first item in the list that satisfies the condition
            subscription = next((item for item in subscriptions if item[1] == associate.associate_id), None)

            if subscription is not None:
                associates_unsubscribed.append({
                    'id': associate.associate_id,
                    'full_name': associate.get_full_name(),
                    'is_minor': associate.is_minor,
                    'subscription_id': subscription[0]
                })

    data = {
        'associates_subscribed': associates_subscribed,
        'associates_unsubscribed': associates_unsubscribed,
        'full': len(associates_unsubscribed) == 0
    }

    return Response(data, status=status.HTTP_200_OK)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def profile_associates_sport_association(request, uid):
    """
    Select all the associates that are subscribed on a sport association and are connected to the user,
    plus return a list of possible subscriptions
    """


    is_valid_uuid(uid)

    user = request.user
    associates = Associate.objects.filter(
        user=user,
        disabled=False
    )
    # get current year subscriptions
    date_from, date_to = BalanceSheetData.get_range_from_year_and_starting_date(
        date=datetime.now(),
        starting_day=user.balance_sheet_start_day,
        starting_month=user.balance_sheet_start_month
    )
    sport_subscriptions = Subscription.objects.filter(
        sport_association_id=uid,
        associate__in=associates,
        creation_date__gte=date_from,
        creation_date__lte=date_to
    )
    associates_subscribed = sport_subscriptions.values_list('associate_id', flat=True)

    # logger.info(associates_subscribed)

    data = {
        'associates_subscribed': [],
        'associates_unsubscribed': []
    }

    for associate in associates:
        if associate.associate_id in associates_subscribed:
            data['associates_subscribed'].append({
                'id': associate.associate_id,
                'full_name': associate.get_full_name(),
                'is_minor': associate.is_minor,
                'subscription_id': sport_subscriptions.filter(associate=associate).first().subscription_id
            })
        else:
            data['associates_unsubscribed'].append({
                'id': associate.associate_id,
                'full_name': associate.get_full_name(),
                'is_minor': associate.is_minor
            })

    return Response(data, status=status.HTTP_200_OK)


@api_view(['GET'])
def profile_image(request, uid):


    is_valid_uuid(uid)

    user = User.objects.get(user_id=uid)
    return Response(user.avatar_image, status=status.HTTP_200_OK)


@api_view(['PATCH'])
@permission_classes([IsAuthenticated])
def profile_update_password(request):
    logger.info("Updating password", extra={'user_id': str(request.user.user_id)})

    password_data = request.data['password_data']
    user = request.original_user if request.collaborator else request.user

    password_pattern = re.compile(r"^(?=.*[A-Z])(?=.*[!@#$&\.\-\_*])(?=.*[0-9]).{10,}$")
    password_validated = re.search(password_pattern, password_data['new_password'])

    content = {}
    if user.check_password(password_data['current_password']) and password_validated:
        logger.info("Password updated successfully", extra={'user_id': str(user.user_id)})
        user.set_password(password_data['new_password'])
        content['msg'] = "password updated with success."
        user.save()
    elif password_validated is None:
        logger.warning("Password update failed - invalid format", extra={'user_id': str(user.user_id)})
        content['msg'] = "password not valid."
        return Response(content, status=status.HTTP_400_BAD_REQUEST)
    else:
        logger.warning("Password update failed - incorrect current password", extra={'user_id': str(user.user_id)})
        content['code'] = ApiMessages.PASSWORD_NOT_VALID
        content['msg'] = "old password not correct."

    # logger.info(content)
    content['role'] = User.ROLE_CHOICES[user.role - 1][1]
    content['user_data'] = UserAuthSerializer(user).data
    # compress avatar image
    if content['user_data']['avatar_image'] is not None:
        content['user_data']['avatar_image'] = compress_base64(content['user_data']['avatar_image'])

    if user.role == User.ASSOCIATION:
        sport_association = SportAssociation.objects.get(user=user)
        content['user_data']['sport_association'] = SportAssociationSerializer(sport_association).data
        # compress logo image
        if content['user_data']['sport_association']['logo'] is not None:
            content['user_data']['sport_association']['logo'] = compress_base64(
                content['user_data']['sport_association']['logo'])

    return Response(content, status=status.HTTP_200_OK)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def profile_settings_tables(request):

    user = UserTablesSettingsSerializer(request.user, data=request.data, partial=True)

    if user.is_valid(raise_exception=True):
        user.save()

    return Response({'status': 'success'}, status=status.HTTP_200_OK)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def profile_invoice_list(request):

    invoices = SportAssociationInvoices.objects.filter(sport_association=request.user.sport_association).order_by(
        '-date_created')
    invoices_list = SportAssociationInvoiceSerializer(invoices, many=True).data

    return Response({'data': invoices_list}, status=status.HTTP_200_OK)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def profile_invoice_download(request, sport_association_invoice_id):

    invoices = SportAssociationInvoices.objects.filter(
        sport_association=request.user.sport_association,
        sport_association_invoice_id=sport_association_invoice_id
    ).first()

    if invoices is None:
        return Response({'error': 'invoice not found'}, status=status.HTTP_404_NOT_FOUND)

    return Response({'data': invoices.invoice}, status=status.HTTP_200_OK)


@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def profile_settings(request):


    user: User = request.user
    membership_card_configuration: SportAssociationMembershipCardConfiguration = SportAssociationMembershipCardConfiguration.objects.filter(sport_association=user.sport_association).first()

    if membership_card_configuration is None:
        membership_card_configuration = SportAssociationMembershipCardConfiguration(sport_association=user.sport_association)
        membership_card_configuration.save()

    if request.method == 'GET':
        date_from, date_to = BalanceSheetData.get_range_from_year_and_starting_date(
            date=datetime.now(),
            starting_day=user.balance_sheet_start_day,
            starting_month=user.balance_sheet_start_month
        )
        return Response({
            "settings": {
                "enumerate_invoices": user.enumerate_invoices,
                "online_payments": user.online_payments,
                "balance_sheet_year": str(user.balance_sheet_year),
                "balance_sheet_start_day": int(user.balance_sheet_start_day),
                "balance_sheet_start_month": int(user.balance_sheet_start_month),
                "temporary_invoice_deletion": user.temporary_invoice_deletion,
                "auto_archive": user.auto_archive,
                "auto_mark_attendance": user.auto_mark_attendance,
                "current_year": date_from.year,
                "payment_date_equal_invoice_date": user.payment_date_equal_invoice_date,
                "starting_number_invoices": user.starting_number_invoices,
                "auto_paid_payment": user.auto_paid_payment,
                "full_installments_plan": user.full_installments_plan,
                "show_zero_payments": user.show_zero_payments,
                "table_settings": user.tables_settings,
                "dark_mode": user.dark_mode,
                "subscription_duration_equal_sport_year": user.subscription_duration_equal_sport_year,
                "subscription_duration": user.subscription_duration,
                "membership_duration": user.membership_duration,
                "membership_starting_number": user.membership_starting_number,
                "default_membership_type": user.default_membership_type,
                "subscription_start_day": user.subscription_start_day,
                "subscription_start_month": user.subscription_start_month,
                "custom_end_date": user.custom_end_date,
                "subscription_end_day": user.subscription_end_day,
                "subscription_end_month": user.subscription_end_month,
                "disable_account_creation": user.disable_account_creation,
                "force_account_creation": user.force_account_creation,
                'medical_certificate_notifications': user.medical_certificate_notifications,
                'hide_category_name': user.hide_category_name,
                'default_payment_category': user.default_payment_category.payment_category_id if user.default_payment_category is not None else None,
                'default_payment_category_courses': user.default_payment_category_courses.payment_category_id if user.default_payment_category_courses is not None else None,
                'membership_card_configuration': {
                    'emit_only_on_approval': membership_card_configuration.emit_only_on_approval,
                    'customized_template': membership_card_configuration.customized_template
                }
            }
        }, status=status.HTTP_200_OK)
    elif request.method == 'POST':
        settings = UserSettingsSerializer(data=request.data)
        # TODO: refactor to update only what is changed via PATCH
        if settings.is_valid(raise_exception=True):
            user.enumerate_invoices = settings.validated_data['enumerate_invoices']
            user.online_payments = settings.validated_data['online_payments']
            user.balance_sheet_year = settings.validated_data['balance_sheet_year']
            # add checks later on for default balance sheet periods
            user.balance_sheet_start_day = settings.validated_data['balance_sheet_start_day']
            user.balance_sheet_start_month = settings.validated_data['balance_sheet_start_month']
            user.temporary_invoice_deletion = settings.validated_data['temporary_invoice_deletion']
            user.auto_archive = settings.validated_data['auto_archive']
            user.auto_mark_attendance = settings.validated_data['auto_mark_attendance']
            user.payment_date_equal_invoice_date = settings.validated_data['payment_date_equal_invoice_date']
            user.starting_number_invoices = settings.validated_data['starting_number_invoices']
            user.auto_paid_payment = settings.validated_data['auto_paid_payment']
            user.full_installments_plan = settings.validated_data['full_installments_plan']
            user.show_zero_payments = settings.validated_data['show_zero_payments']
            user.dark_mode = settings.validated_data['dark_mode']
            user.medical_certificate_notifications = settings.validated_data['medical_certificate_notifications']
            user.hide_category_name = settings.validated_data['hide_category_name']
            user.subscription_duration_equal_sport_year = settings.validated_data['subscription_duration_equal_sport_year']
            user.disable_account_creation = settings.validated_data['disable_account_creation']
            user.force_account_creation = settings.validated_data['force_account_creation']
            user.subscription_duration = settings.validated_data['subscription_duration']
            user.membership_duration = settings.validated_data['membership_duration']
            user.membership_starting_number = settings.validated_data['membership_starting_number']
            user.default_membership_type = settings.validated_data['default_membership_type']
            user.subscription_start_day = settings.validated_data['subscription_start_day']
            user.subscription_start_month = settings.validated_data['subscription_start_month']
            user.custom_end_date = settings.validated_data['custom_end_date']
            user.subscription_end_day = settings.validated_data['subscription_end_day']
            user.subscription_end_month = settings.validated_data['subscription_end_month']
            user.default_payment_category = settings.validated_data['default_payment_category']
            user.default_payment_category_courses = settings.validated_data['default_payment_category_courses']
            user.save()
        # get membership card configuration
        membership_card_configuration = SportAssociationMembershipCardConfiguration.objects.filter(sport_association=user.sport_association).first()
        if membership_card_configuration is None:
            return Response({'status': 'membership card configuration not found'}, status=status.HTTP_404_NOT_FOUND)
        else:
            membership_card_configuration.emit_only_on_approval = request.data['membership_card_configuration']['emit_only_on_approval']
            membership_card_configuration.customized_template = request.data['membership_card_configuration']['customized_template']
            membership_card_configuration.save()
        return Response({'status': 'success'}, status=status.HTTP_200_OK)
    else:
        return Response(status=status.HTTP_401_UNAUTHORIZED)


@api_view(['GET', 'PATCH'])
@permission_classes([IsAuthenticated])
def profile_integrations(request):

    if request.method == 'GET':
        sport_association = request.user.sport_association
        return Response({
            'review_url': sport_association.review_url,
            'review_url_enabled': sport_association.review_url_enabled
        }, status=status.HTTP_200_OK)

    if request.method == 'PATCH':
        sport_association = request.user.sport_association
        sport_association.review_url = request.data.get('review_url', sport_association.review_url)
        sport_association.review_url_enabled = request.data.get('review_url_enabled', sport_association.review_url_enabled)
        sport_association.save()
        return Response({
            'review_url': sport_association.review_url,
            'review_url_enabled': sport_association.review_url_enabled
        }, status=status.HTTP_200_OK)

    return Response(status=status.HTTP_405_METHOD_NOT_ALLOWED)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def sport_association_list(request):
    """
    List all sport associations
    """
    # check if superuser
    if not request.user.is_superuser:
        return Response(status=status.HTTP_401_UNAUTHORIZED)
    sport_associations = SportAssociation.objects.all().iterator(chunk_size=100)
    serializer = SportAssociationToolSerializer(sport_associations, many=True)
    return Response(serializer.data)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def sport_association_admin_update(request, uid):
    """
    List all sport associations
    """
    # check if superuser
    if not request.user.is_superuser:
        return Response(status=status.HTTP_401_UNAUTHORIZED)

    is_valid_uuid(uid)
    sport_association = SportAssociation.objects.filter(sport_association_id=uid).first()

    if sport_association is None:
        return Response(status=status.HTTP_404_NOT_FOUND)

    billing_subscription = BillingSubscription.objects.filter(
        user=sport_association.user
    ).first()

    if billing_subscription is None:
        return Response(status=status.HTTP_404_NOT_FOUND)

    # update the billing subscription from post data
    billing_data = request.data.get('billing_subscription', None)
    if billing_data is None:
        return Response(status=status.HTTP_400_BAD_REQUEST)
    billing_subscription.auto_renewal = billing_data.get('auto_renewal', billing_subscription.auto_renewal)
    billing_subscription.renewal_type = billing_data.get('renewal_type', billing_subscription.renewal_type)
    billing_subscription.ends_on = billing_data.get('ends_on', billing_subscription.ends_on)
    billing_subscription.billing_plan_id = billing_data.get('billing_plan', billing_subscription.billing_plan_id)
    billing_subscription.save()
    
    # notes 
    try:
        sport_association.notes = request.data.get('notes', sport_association.notes)
        sport_association.save()
    except Exception as e:
        logger.error(f"Error updating notes: {e}")
        return Response({"msg": "Errore durante l'aggiornamento delle note."}, status=status.HTTP_400_BAD_REQUEST)
    
    return Response({"msg": "Modifiche salvate."}, status=status.HTTP_200_OK)




@api_view(['GET'])
@permission_classes([IsAuthenticated])
# @cache_endpoint('info', timeout=60)
def profile_info(request):


    user = request.user

    content = {}
    content['user_data'] = UserAuthSerializer(user).data
    # need to compress the avatar image (which is a base64 string) to save bandwidth and improve performance
    if content['user_data']['avatar_image'] is not None:
        content['user_data']['avatar_image'] = compress_base64(content['user_data']['avatar_image'])
    if user.role == User.ASSOCIATION or user.role == User.COLLABORATOR:
        if user.role == User.ASSOCIATION:
            sport_association = SportAssociation.objects.get(user=user)
        else:
            sport_association = SportAssociation.objects.get(user=user.connected_user)
        content['user_data']['sport_association'] = SportAssociationSerializer(sport_association).data
        # need to compress the logo (which is a base64 string) to save bandwidth and improve performance
        if content['user_data']['sport_association']['logo'] is not None:
            content['user_data']['sport_association']['logo'] = compress_base64(
                content['user_data']['sport_association']['logo']
            )
        content['user_data']['temporary_invoice_deletion'] = user.temporary_invoice_deletion
        if sport_association.regulation is None or sport_association.regulation == '' or \
                sport_association.demand is None or sport_association.demand == '':
            content['user_data']['sport_association']['empty_sections'] = True
        content['user_data']['sport_association']['review_url'] = sport_association.review_url
        content['user_data']['sport_association']['review_url_enabled'] = sport_association.review_url_enabled
        content['user_data']['preview_and_custom_features'] = user.get_preveiw_and_custom_features()
    # check if superuser
    if user.is_superuser:
        content['user_data']['is_superuser'] = True

    return Response({
        "info": {
            "role": User.ROLE_CHOICES[user.role-1][1],
        },
        "user_data": content['user_data']
    }, status=status.HTTP_200_OK)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def testimonials_create(request):

    data = request.data

    if not request.user.is_sport_association:
        return Response({'msg': 'not a sport association.'}, status=status.HTTP_401_UNAUTHORIZED)

    if 'text' not in data or 'score' not in data:
        return Response({'msg': 'missing params.'}, status=status.HTTP_400_BAD_REQUEST)

    if data['score'] < 1 or data['score'] > 5:
        return Response({'msg': 'invalid score.'}, status=status.HTTP_400_BAD_REQUEST)

    # clean text from html tags and injection attacks
    # data['text'] = BeautifulSoup(data['text']).prettify()

    # add testimonial to db
    testimonial = Testimonial.objects.create(
        sport_association=request.user.sport_association,
        text=data['text'],
        score=data['score']
    )
    testimonial.save()

    sport_association = SportAssociation.objects.filter(user=request.user)
    if sport_association is None:
        return Response({'msg': 'sport association not found.'}, status=status.HTTP_404_NOT_FOUND)
    else:
        sport_association = sport_association.first()
        sport_association.reviewed = True
        sport_association.save()

    return Response({'msg': 'Thank you for your review.'}, status=status.HTTP_200_OK)


@api_view(['POST'])
def testimonials_update(request):
    data = request.data

    # logger.info(f"request from senja with data: {data}")
    # logger.info(f"headers are: {request.META}")
    # check if there is header HTTP_X_SENJA_SIGNATURE
    if 'x-senja-signature' not in request.headers:
        return Response(status=status.HTTP_401_UNAUTHORIZED)

    # check if signature is valid
    payload_bytes = json.dumps(data, separators=(',', ':')).encode()
    signature = hmac.new(settings.SENJA_WEBHOOK_SECRET.encode(), payload_bytes, hashlib.sha256).hexdigest()

    if not hmac.compare_digest(signature, request.headers['x-senja-signature']):
        return Response(status=status.HTTP_401_UNAUTHORIZED)

    # extract the type of request
    if 'type' not in data:
        return Response(status=status.HTTP_400_BAD_REQUEST)

    if data['type'] == 'testimonial_created':
        try:
            sport_association_id = data['data']['new']['endorser']['custom_data']['sport-association-id']

            # check if sport association id is valid
            is_valid_uuid(sport_association_id)

            sport_association = SportAssociation.objects.filter(
                sport_association_id=sport_association_id).first()
            if sport_association is None:
                return Response(status=status.HTTP_404_NOT_FOUND)

            sport_association.reviewed = True
            sport_association.save()
        except ValidationError as e:
            logger.error(e)
            return Response(status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            logger.error(e)
            return Response(status=status.HTTP_400_BAD_REQUEST)

    return Response(status=status.HTTP_200_OK)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def export_all_data(request):

    user = request.user

    if not user.is_sport_association:
        return Response({'msg': 'not a sport association.'}, status=status.HTTP_401_UNAUTHORIZED)


    wb = Workbook()
    # Remove the default sheet created
    wb.remove(wb.active)

    # create a new sheet Iscrizioni
    subscriptions_sheet = wb.create_sheet(title='Iscrizioni')

    # perform raw query to get all the subscriptions
    subscriptions = Subscription.objects.filter(
        sport_association=user.sport_association
    ).select_related('associate', 'medical').prefetch_related('associate__tutors')
    # add all columns to the sheet dynamically, also of the associate and associate__tutors that is ManyToMany
    # Get field names for each model
    subscription_fields = [str(f.name) for f in Subscription._meta.fields]
    medical_fields = [str(f.name) for f in MedicalCertificate._meta.fields]
    associate_fields = [str(f.name) for f in Associate._meta.fields]
    tutor_fields = [str(f.name) for f in Associate._meta.fields]

    # exclude the headers that are not in this list
    EXCLUDE_HEADERS = [
        'associate__tutors',
        'signature',
        'associate',
        'supplier',
        'instructor',
        'medical',
        'medical_id',
        'sport_association',
        'associate__sport_association',
        'tutor__sport_association',
        'tutor__associate_id',
        'associate__associate_id',
        'associate_id',
        'payment',
        'document',
        'subscription_id',
        'custom_data',
        'additional_fields',
        'user',
        'associate__user',
        'tutor__user',
        'deleted',
        'trial',
        'document_pdf',
        'associate__picture_path',
        'picture_path',
        'meta',
    ]

    # exclude the headers that are not in this list
    subscription_fields = [field for field in subscription_fields if field not in EXCLUDE_HEADERS]
    medical_fields = [field for field in medical_fields if field not in EXCLUDE_HEADERS]
    associate_fields = [field for field in associate_fields if field not in EXCLUDE_HEADERS]
    tutor_fields = [field for field in tutor_fields if field not in EXCLUDE_HEADERS]

    # Create headers
    headers = (
            subscription_fields +
            medical_fields +
            [f'associate__{field}' for field in associate_fields] +
            [f'tutor__{field}' for field in tutor_fields]
    )


    subscriptions_sheet.append(headers)

    # Add data rows
    for subscription in subscriptions:
        row = []

        # Add subscription fields
        for field in subscription_fields:
            if field in EXCLUDE_HEADERS:
                continue
            if field == 'status_flag':
                row.append(str(Subscription.STATUS_FLAG[subscription.status_flag - 1][1]))
            elif field == 'type':
                row.append(str(Subscription.TYPE[subscription.type - 1][1]))
            elif field == 'role':
                row.append(str(Subscription.ASSOCIATE_ROLE[subscription.role - 1][1]))
            else:
                row.append(str(getattr(subscription, field)) if getattr(subscription, field) is not None else '-')

        # Add medical fields
        for field in medical_fields:
            if field in EXCLUDE_HEADERS:
                continue
            if subscription.medical is None:
                row.append('-')
            else:
                row.append(str(getattr(subscription.medical, field) if getattr(subscription.medical, field) is not None else '-'))

        # Add associate fields
        for field in associate_fields:
            if field in EXCLUDE_HEADERS:
                continue
            row.append(str(getattr(subscription.associate, field) if getattr(subscription.associate, field) is not None else '-'))

        # Add tutor fields - handle multiple tutors
        tutors = subscription.associate.tutors.all()
        tutor_data = []
        for tutor in tutors:
            for field in tutor_fields:
                if field in EXCLUDE_HEADERS:
                    continue
                tutor_data.append(str(getattr(tutor, field) if getattr(tutor, field) is not None else '-'))
        row.extend(tutor_data)

        subscriptions_sheet.append(row)

    # create a new sheet Pagamenti
    payments_sheet = wb.create_sheet(title='Pagamenti')

    # perform raw query to get all the payments
    payments = Payment.objects.filter(
        sport_association=user.sport_association
    ).select_related('associate', 'payment_category', 'invoice', 'supplier', 'instructor', 'custom_accounts')

    # Get field names for each model
    payment_fields = [str(f.name) for f in Payment._meta.fields]
    payment_category_fields = [str(f.name) for f in PaymentCategory._meta.fields]
    custom_accounts_fields = [str(f.name) for f in CustomAccounts._meta.fields]
    associate_fields = [str(f.name) for f in Associate._meta.fields]
    supplier_fields = [str(f.name) for f in SupplierAndCustomers._meta.fields]
    instructor_fields = [str(f.name) for f in Instructor._meta.fields]
    invoice_fields = [str(f.name) for f in Invoice._meta.fields]

    # exclude the headers that are not in this list
    PAYMENT_EXCLUDE_HEADERS = [
        'associate',
        'instructor',
        'supplier',
        'payment_category',
        'payment_id',
        'invoice',
        'user',
        'sport_association',
        'payment_intent_id',
        'invoice_fiscal_code',
        'custom_accounts',
        'custom_account_id',
        'deleted',
        'enabled',
        'editable',
        'course_id',
        'associate_id',
        'supplier_id',
        'instructor_id',
        'document_pdf',
        'imported_from_associami',
        'picture_path',
        'payment_category_id',
        'invoice_id',
    ]

    # exclude the headers that are not in this list
    payment_fields = [field for field in payment_fields if field not in PAYMENT_EXCLUDE_HEADERS]
    payment_category_fields = [field for field in payment_category_fields if field not in PAYMENT_EXCLUDE_HEADERS]
    custom_accounts_fields = [field for field in custom_accounts_fields if field not in PAYMENT_EXCLUDE_HEADERS]
    supplier_fields = [field for field in supplier_fields if field not in PAYMENT_EXCLUDE_HEADERS]
    associate_fields = [field for field in associate_fields if field not in PAYMENT_EXCLUDE_HEADERS]
    instructor_fields = [field for field in instructor_fields if field not in PAYMENT_EXCLUDE_HEADERS]
    invoice_fields = [field for field in invoice_fields if field not in PAYMENT_EXCLUDE_HEADERS]

    # Create headers
    headers = (
            payment_fields +
            ['course_name'] +
            [f'payment_category__{field}' for field in payment_category_fields] +
            [f'custom_accounts__{field}' for field in custom_accounts_fields] +
            [f'associate__{field}' for field in associate_fields] +
            [f'supplier__{field}' for field in supplier_fields] +
            [f'instructor__{field}' for field in instructor_fields] +
            [f'invoice__{field}' for field in invoice_fields]
    )

    payments_sheet.append(headers)

    def get_course(obj):
        if obj.subject == Payment.COURSE:
            try:
                sub = CourseSubscription.objects.get(payment=obj)
                return sub.course.title
            except CourseSubscription.DoesNotExist:
                try:
                    sub = CourseSubscriptionInstallment.objects.get(payment=obj)
                    return sub.course_subscription.course.title
                except CourseSubscriptionInstallment.DoesNotExist:
                    if obj.course:
                        return obj.course['label'] if 'label' in obj.course else None
                    return None
        else:
            return None

    # Add data rows
    for payment in payments:
        row = []

        # Add payment fields
        for field in payment_fields:
            if field in PAYMENT_EXCLUDE_HEADERS:
                continue
            if field == 'subject':
                row.append(str(Payment.SUBJECT[payment.subject][1]))
            elif field == 'type':
                row.append(dict(Payment.TYPE_CHOICES).get(payment.type, '-'))
            else:
                row.append(str(getattr(payment, field)) if getattr(payment, field) is not None else '-')

        # Add course name
        course_name = get_course(payment)
        if course_name is None:
            row.append('-')
        else:
            row.append(str(course_name))

        # Add payment category fields
        for field in payment_category_fields:
            if field in PAYMENT_EXCLUDE_HEADERS:
                continue
            if payment.payment_category is None:
                row.append('-')
            else:
                row.append(str(getattr(payment.payment_category, field) if getattr(payment.payment_category, field) is not None else '-')
                           )

        # Add custom accounts fields
        for field in custom_accounts_fields:
            if field in PAYMENT_EXCLUDE_HEADERS:
                continue
            if payment.custom_accounts is None:
                row.append('-')
            else:
                if field == 'account_type':
                    row.append(str(CustomAccounts.ACCOUNT_TYPE[payment.custom_accounts.account_type - 1][1]))
                else:
                    row.append(str(getattr(payment.custom_accounts, field) if getattr(payment.custom_accounts, field) is not None else '-')
                               )

        # Add associate fields
        for field in associate_fields:
            if field in PAYMENT_EXCLUDE_HEADERS:
                continue
            if payment.associate is None:
                row.append('-')
            else:
                row.append(str(getattr(payment.associate, field) if getattr(payment.associate, field) is not None else '-')
                           )

        # Add supplier fields
        for field in supplier_fields:
            if field in PAYMENT_EXCLUDE_HEADERS:
                continue
            if payment.supplier is None:
                row.append('-')
            else:
                row.append(str(getattr(payment.supplier, field) if getattr(payment.supplier, field) is not None else '-')
                           )

        # Add instructor fields
        for field in instructor_fields:
            if field in PAYMENT_EXCLUDE_HEADERS:
                continue
            if payment.instructor is None:
                row.append('-')
            else:
                row.append(str(getattr(payment.instructor, field) if getattr(payment.instructor, field) is not None else '-')
                           )

        # Add invoice fields
        for field in invoice_fields:
            if field in PAYMENT_EXCLUDE_HEADERS:
                continue
            if payment.invoice is None:
                row.append('-')
            else:
                row.append(str(getattr(payment.invoice, field) if getattr(payment.invoice, field) is not None else '-')
                           )

        payments_sheet.append(row)

    # Prepare the downloadable file response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=export_{user.sport_association.denomination}.xlsx'

    # Save the workbook to the response
    wb.save(response)
    return response
